import customtkinter as ctk
import tkinter.messagebox as tkmb
import tkinter as tk
from tkinter import filedialog, scrolledtext
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import pyperclip
import os
import mysql.connector
import pandas as pd
from test import dataRefine,Database

# Set custom appearance and color theme
ctk.set_appearance_mode("light")  # The custom color theme will be applied manually

# Initialize main application
class App():
    def __init__(self):
        self.app = ctk.CTk()
        self.app.geometry(f"{self.app.winfo_screenwidth()}x{self.app.winfo_screenheight()}")
        self.app.title("Salary-slip Generator")
    
        self.children = {'login':self.Login(self,self.app),'fileinput':self.FileInput(self,self.app),'interface':self.Interface(self,self.app)}
        self.children['login'].appear()

    class Login():
        def __init__(self,outer,master):
            self.visible = False
            self.outer = outer
            self.frame = ctk.CTkScrollableFrame(master=master, fg_color=custom_color_scheme["fg_color"])

            self.label_login = ctk.CTkLabel(master=self.frame, text='Admin login page', text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.label_login.pack(pady=20)

            self.user_entry = ctk.CTkEntry(master=self.frame, placeholder_text="Username", text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.user_entry.pack(pady=12, padx=10)

            self.user_pass = ctk.CTkEntry(master=self.frame, placeholder_text="Password", show="*", text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.user_pass.pack(pady=12, padx=10)

            self.button_login = ctk.CTkButton(master=self.frame, text='Login', command=self.login, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_login.pack(pady=12, padx=10)

            self.checkbox = ctk.CTkCheckBox(master=self.frame, text='Remember Me', fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.checkbox.pack(pady=12, padx=10)


        def appear(self):
            for child in self.outer.children:
                self.outer.children[child].hide()
                
            if not self.visible:
                self.frame.pack(pady=20, padx=40, fill='both', expand=True)
            else:
                print(' Already visible')

        def hide(self):
            if self.visible:
                self.frame.pack_forget()
            else:
                print(' Already         ')

        def login(self):
            known_user = 'admin'
            known_pass = 'kjs2024'
            username = self.user_entry.get()
            password = self.user_pass.get()

            if known_user == username and known_pass == password:
                tkmb.showinfo(title="Login Successful", message="You have logged in Successfully")
                show_input_page()
            elif known_user == username and known_pass != password:
                tkmb.showwarning(title='Wrong password', message='Please check your password')
            elif known_user != username and known_pass == password:
                tkmb.showwarning(title='Wrong username', message='Please check your username')
            else:
                tkmb.showerror(title="Login Failed", message="Invalid Username and password")

    class Interface():
        def __init__(self,outer,master):
            self.visible = False

            self.frame = ctk.CTkScrollableFrame(master=master, fg_color=custom_color_scheme["fg_color"])

            self.label_input = ctk.CTkLabel(master=self.frame , text="Enter Details", text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.label_input.pack(pady=20)

            self.entry_month = ctk.CTkEntry(master=self.frame , placeholder_text="Month (e.g., May)", text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.entry_month.pack(pady=12, padx=10)

            self.entry_year = ctk.CTkEntry(master=self.frame , placeholder_text="Year (e.g., 2024)", text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.entry_year.pack(pady=12, padx=10)

            self.toggle_institute = ctk.CTkComboBox(master=self.frame , values=["Somaiya", "SVV"], fg_color=custom_color_scheme["combo_box_color"], font=("Helvetica", 16))
            self.toggle_institute.pack(pady=12, padx=10)

            self.button_continue = ctk.CTkButton(master=self.frame , text='Continue', command=check_excel_file, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_continue.pack(pady=12, padx=10)

            self.button_view_db = ctk.CTkButton(master=self.frame , text='View Data from DB', command=fetchDatabase, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_view_db.pack(pady=12, padx=10)

            self.button_back_to_login = ctk.CTkButton(master=self.frame , text='Back', command=show_login_page, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_back_to_login.pack(pady=12, padx=10)


        def appear(self):
            for child in self.outer.children:
                self.outer.children[child].hide()
                
            if not self.visible:
                self.frame.pack(pady=20, padx=40, fill='both', expand=True)
            else:
                print(' Already visible')

        def hide(self):
            if self.visible:
                self.frame.pack_forget()
            else:
                print(' Already         ')

    class FileInput():
        def __init__(self,outer,master):
            self.visible = False

            self.frame = ctk.CTkScrollableFrame(master=master, fg_color=custom_color_scheme["fg_color"])

            self.label_file = ctk.CTkLabel(master=self.frame , text="Selected Excel File:", text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.label_file.pack(pady=10)

            self.entry_file = ctk.CTkEntry(master=self.frame , width=50, text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.entry_file.pack(pady=5)

            self.button_browse = ctk.CTkButton(master=self.frame , text="Browse", command=select_file, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_browse.pack(pady=5)

            self.label_id = ctk.CTkLabel(master=self.frame , text="Enter Employee ID:", text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.label_id.pack()

            self.entry_id = ctk.CTkEntry(master=self.frame , text_color=custom_color_scheme["text_color"], font=("Helvetica", 16))
            self.entry_id.pack(pady=5)

            self.button_extract = ctk.CTkButton(master=self.frame , text="Generate pdf", command=extract_data, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_extract.pack(pady=10)

            self.button_bulk_print = ctk.CTkButton(master=self.frame , text="Bulk Print PDFs", command=bulk_print_pdfs, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_bulk_print.pack(pady=10)

            self.button_copy = ctk.CTkButton(master=self.frame , text="Copy Row to Clipboard", command=copy_row_to_clipboard, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_copy.pack(pady=10)

            self.button_back_to_input = ctk.CTkButton(master=self.frame , text='Back', command=input, fg_color=custom_color_scheme["button_color"], font=("Helvetica", 16))
            self.button_back_to_input.pack(pady=12, padx=10)

            self.text_excel = scrolledtext.ScrolledText(master=self.frame , width=90, height=25, bg="black", fg="white", wrap=tk.NONE, font=("Courier", 12))
            self.text_excel.pack(pady=10, fill='both', expand=True)

        def appear(self):
            for child in self.outer.children:
                self.outer.children[child].hide()
                
            if not self.visible:
                self.frame.pack(pady=20, padx=40, fill='both', expand=True)
            else:
                print(' Already visible')

        def hide(self):
            if self.visible:
                self.frame.pack_forget()
            else:
                print(' Already         ')


    


def fetchDatabase():
    pde = pd.read_excel('front/KJSIT_MAY_2023.xlsx')
    c = dataRefine(pde)
    new_col = c.refine()
    b = Database(
        host="localhost",
        user="root",
        password="1234",
        database="somaiya_salary", columns=new_col.columns
    )


  
def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", ".xlsx;.xls")])
    if file_path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_path)
        view_excel(file_path)

def view_excel(file_path):
    text_excel.delete(1.0, tk.END)
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows():
        row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
        data.append(row_data)

    col_widths = [max(len(str(cell)) for cell in col) for col in zip(*data)]
    formatted_data = "\n".join(" | ".join(f"{cell:<{col_widths[i]}}" for i, cell in enumerate(row)) for row in data)
    
    text_excel.insert(tk.END, formatted_data)

def extract_data():
    file_path = entry_file.get()
    employee_id = entry_id.get()
    if not file_path:
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == employee_id:
            employee_data = [cell.value for cell in row]
            generate_pdf(employee_data)
            break
    else:
        tkmb.showwarning("Error", "Employee ID not found.")

def generate_pdf(employee_data):
    try:
        pdf_file = f"employee_{employee_data[0]}.pdf"
        doc = SimpleDocTemplate(pdf_file, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        style_normal = styles["Normal"]
        style_highlight = styles["BodyText"]

        # Title
        title_data = [
            ["K.J SOMAIYA INSTITUTE OF ENGINEERING & INFORMATION TECHNOLOGY, SOMAIYA AYURVIHAR EVARAD NAGAR, EASTERN EXPRESS HIGHWAY SION"],
            [f"PAY SLIP FOR THE MONTH OF {entry_month.get()}-{entry_year.get()}     31 DAYS     1"]
        ]
        title_table = Table(title_data, colWidths=[540])
        title_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (0, 0)),
            ('SPAN', (0, 1), (0, 1)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold')
        ]))
        elements.append(title_table)
        elements.append(Paragraph("<br/><br/>", style_normal))  # Add spacing

        # Employee Info
        employee_info_data = [
            ["NAME", employee_data[1], "DESIGNATION", employee_data[2]],
            ["DATE OF JOINING", employee_data[3], "NO OF DAYS PRESENT", employee_data[4]],
            ["PF NO", employee_data[5], "PAN NO", employee_data[6]],
            ["EMP CODE", employee_data[0], "SALARY A/C NO", employee_data[7]],
            ["Aadhar Card No", employee_data[8], "UNA", employee_data[9]],
        ]
        employee_info_table = Table(employee_info_data, colWidths=[120, 160, 120, 160])
        employee_info_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold')
        ]))
        elements.append(employee_info_table)
        elements.append(Paragraph("<br/><br/>", style_normal))  # Add spacing

        # Earnings and Deductions
        earnings_deductions_data = [
            ["EARNINGS", "RS", "DEDUCTIONS", "RS"],
            ["Basic Pay", employee_data[10], "PROF TAX", employee_data[26]],
            ["Basic", employee_data[10], "", ""],
            ["RENT HRA", employee_data[11], "PF", employee_data[27]],
            ["30%", employee_data[11], "TDS", employee_data[28]],
            ["TA/ Conveyance", employee_data[12], "LIC", employee_data[29]],
            ["SPECIAL ALW", employee_data[13], ""],
            ["Salary Arrears", employee_data[14], "Principle loan amount PM", employee_data[30]],
            ["Vehicle", employee_data[15], "", ""],
            ["Books and Periodicals", employee_data[16], "Interest 6% on bal amount", employee_data[30]],
            ["Etn Alw", employee_data[17], "", ""],
            ["Petrol Alw", employee_data[18], "", ""],
            ["Telephone", employee_data[19], "Other Deduction", employee_data[30]],
            ["Medical", employee_data[20], "", ""],
            ["LTA", employee_data[21], "", ""],
            ["Alw", employee_data[22], "", ""],
            ["EX-Grataia", employee_data[23], "", ""],
            ["Ent All", employee_data[24], "", ""],
            ["GROSS SALARY", employee_data[25], "TOTAL DEDUCTION", employee_data[30]],
            ["NET SALARY PAYABLE", employee_data[30], "", ""]
        ]
        earnings_deductions_table = Table(earnings_deductions_data, colWidths=[180, 100, 180, 100])
        earnings_deductions_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER')
        ]))
        elements.append(earnings_deductions_table)
        elements.append(Paragraph("<br/><br/>", style_normal))  # Add spacing

        # Footer
        footer_data = [
            ["This is a computer generated salary slip", ""],
            ["ACCOUNT OFFICER", ""]
        ]
        footer_table = Table(footer_data, colWidths=[300, 240])
        footer_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (-1, 0)),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (0, 1), (0, 1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold')
        ]))
        elements.append(footer_table)

        doc.build(elements)
        #tkmb.showinfo("PDF Generated", f"PDF generated: {pdf_file}")
    except Exception as e:
        tkmb.showerror("Error", f"An error occurred while generating the PDF: {str(e)}")



def copy_row_to_clipboard():
    file_path = entry_file.get()
    employee_id = entry_id.get()
    if not file_path:
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == employee_id:
            employee_data = [str(cell.value) for cell in row]
            row_data = "\t".join(employee_data)
            pyperclip.copy(row_data)
            tkmb.showinfo("Copy Row", "Employee data copied to clipboard.")
            break
    else:
        tkmb.showwarning("Error", "Employee ID not found.")

def bulk_print_pdfs():
    file_path = entry_file.get()
    if not file_path:
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        employee_data = [cell.value for cell in row]
        generate_pdf(employee_data)
    
    tkmb.showinfo("Bulk Print", "Bulk PDF generation completed.")

def check_excel_file():
    month = entry_month.get()
    year = entry_year.get()
    institute = toggle_institute.get()
    
    file_name = f"{institute}{month}{year}.xlsx"
    if os.path.exists(file_name):
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_name)
        tkmb.showinfo("File Found", "The specified Excel file is found.")
    else:
        tkmb.showwarning("File Not Found", "The specified Excel file is not found. Please upload the required file.")
    
    show_file_view_page()


def show_file_view_page():
    frame_input.pack_forget()
    frame_view.pack(pady=20, padx=40, fill='both', expand=True)

def show_login_page():
    frame_input.pack_forget()
    frame_view.pack_forget()
    frame_login.pack(pady=20, padx=40, fill='both', expand=True)

def show_input_page():
    frame_login.pack_forget()
    frame_view.pack_forget()
    frame_input.pack(pady=20, padx=40, fill='both', expand=True)

# Apply custom colors
custom_color_scheme = {
    "fg_color": "white",
    "button_color": "dark red",
    "text_color": "black",
    "combo_box_color": "white"  # Added for combo box color
}


app = App()

app.app.mainloop()
