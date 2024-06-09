import openpyxl
import random

# Load the existing Excel file
file_path = 'Sample sheet for salary calculation and salary slip (1).xlsx'
workbook = openpyxl.load_workbook(file_path)

# Add a new worksheet for random pay slip data
work = workbook.create_sheet('Pay Slips')

# Define the header
header = [
    'NAME', 'DATE OF JOINING', 'PF NO', 'EMP CODE', 'AADHAR CARD NO', 'DESIGNATION',
    'NO OF DAYS PRESENT', 'PAN NO', 'SALARY A/C NO', 'UNA NO',
    'BASIC PAY', 'PRESENT BASIC', 'DP /GP', 'DA', 'HRA', 'CLA', 'TA',
    'SPECIAL ALLW/ME INCERMENT', 'BASIC/DA ARRERS', 'OTHER', 'GROSS SALARY', 'PROF TAX',
    'PF', 'TDS', 'LIC', 'PERSONAL LOAN', 'DIWALI ADVANCE/OTHER',
    'STAFF LOAN/ADVANCE', 'ADVANCE / MEDICAL', 'Emp Credit Soc', 'TOTAL DEDUCTION', 'NET SALARY PAYABLE'
]

# Write the header to the worksheet
for i, j in enumerate(header):
    work.cell(row=1, column=i+1, value=j)

# Generate random data for 26 employees
for i in range(26):
    data = [
        f'Employee {i+1}', f'2021-0{i%12+1}-01', f'MH/28{i:04d}/', f'EMP{i:04d}', f'{random.randint(1000, 9999)}{random.randint(1000, 9999)}',
        f'Designation {i+1}', random.randint(20, 31), f'PAN{i:04d}', f'SALARY_AC_{i:04d}', f'UNA{i:04d}',
        random.randint(10000, 50000), random.randint(10000, 50000), random.randint(1000, 5000), random.randint(1000, 5000),
        random.randint(1000, 5000), random.randint(1000, 5000), random.randint(1000, 5000), random.randint(1000, 5000),
        random.randint(1000, 5000), random.randint(1000, 5000), random.randint(10000, 50000), random.randint(100, 1000),
        random.randint(100, 1000), random.randint(100, 1000), random.randint(100, 1000), random.randint(100, 1000),
        random.randint(100, 1000), random.randint(100, 1000), random.randint(100, 1000), random.randint(100, 1000),
        random.randint(1000, 5000), random.randint(5000, 40000)
    ]
    
    # Write the data to the worksheet
    for j in range(len(data)):
        work.cell(row=i+2, column=j+1, value=data[j])

# Save the modified Excel file
output_path = 'modified_sample_with_payslip.xlsx'
workbook.save(output_path)

print(f"Modified Excel file saved as {output_path}")
